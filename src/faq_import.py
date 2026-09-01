"""FAQ import (Tika unified text extraction + Q/A/S/C rule matching) + downloadable templates.

解析侧改造说明：
  * 所有格式（json/csv/txt/xls/xlsx/pdf/doc/docx/...）**统一**先经 Apache Tika 抽取纯文本；
  * 再按 ``_parse_text_content`` 中的 ``Q:/A:/S:/C:`` 规则匹配 FAQ 条目；
  * 不再依赖 ``json`` / ``csv.DictReader`` / ``openpyxl`` / ``xlrd`` 的专用结构化解析。
模板生成侧保持不变：``_render_*`` 仍用 openpyxl / xlwt / csv / json 标准库生成下载模板，
    这些是生成能力而非解析能力，与 Tika 并不冲突。
"""

from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from src.config import Settings
from src.tika_parser import extract_text_from_bytes

# 表格模板使用的中文表头（CSV / Excel）；导入时由 _norm_header 映射回内部字段名
TABLE_HEADERS_ZH: list[str] = ["编号", "分类", "标准问", "答案", "相似问"]

FIELD_DOCS: list[dict[str, str]] = [
    {
        "field": "id",
        "zh": "编号",
        "required": "否",
        "meaning": (
            "可留空。当前导入不会使用该列，系统会为每条 FAQ 自动生成编号；"
            "模板里填写仅作示例参考。"
        ),
    },
    {
        "field": "category",
        "zh": "分类",
        "required": "否",
        "meaning": "业务分类，便于管理，例如「订单物流」「支付」。",
    },
    {
        "field": "question",
        "zh": "标准问",
        "required": "是",
        "meaning": "该条 FAQ 的主问题表述（用户最常问的那一句）。",
    },
    {
        "field": "answer",
        "zh": "答案",
        "required": "是",
        "meaning": "标准答案正文，作为客服回复依据。",
    },
    {
        "field": "similar",
        "zh": "相似问",
        "required": "否",
        "meaning": (
            "用户换一种说法时的相似问。"
            "JSON 中为字符串数组；表格中多个相似问用 | 分隔，例如：收货地址能改吗|订单地址怎么改"
        ),
    },
]

SAMPLE_ENTRIES: list[dict[str, Any]] = [
    {
        "id": "sample_1",
        "category": "订单物流",
        "question": "如何修改收货地址？",
        "similar": ["收货地址能改吗", "订单地址怎么改"],
        "answer": "未发货订单可在「我的订单」中修改；已发货请联系客服或拒收后重新下单。",
    },
    {
        "id": "sample_2",
        "category": "支付",
        "question": "支持哪些支付方式？",
        "similar": ["可以怎么付款", "支持微信支付吗"],
        "answer": "支持微信、支付宝、银联及企业对公转账（企业版）。",
    },
]

# 模板仍提供 5 种传统格式下载，但实际导入已通过 Tika 统一，以下后缀均可识别
TEMPLATE_FORMATS = ("json", "csv", "txt", "xls", "xlsx")
SUPPORTED_FORMATS = TEMPLATE_FORMATS + (
    "pdf", "doc", "docx", "rtf", "odt",
    "xlsx", "xls",
    "pptx", "ppt", "odp",
    "html", "htm", "eml", "md", "markdown",
)
SUPPORTED_FORMATS = tuple(dict.fromkeys(SUPPORTED_FORMATS))  # 去重保序

_FORMAT_META: dict[str, dict[str, str]] = {
    "json": {
        "filename": "faq_template.json",
        "description": (
            "JSON 模板（下载后请按 Q:/A: 规则填内容）；"
            "导入时由 Apache Tika 抽取纯文本后按 Q/A/S/C 规则识别 FAQ"
        ),
        "media_type": "application/json; charset=utf-8",
    },
    "csv": {
        "filename": "faq_template.csv",
        "description": (
            "CSV 模板（下载后按 Q:/A: 填内容）；"
            "导入时由 Apache Tika 抽取纯文本后统一识别"
        ),
        "media_type": "text/csv; charset=utf-8",
    },
    "txt": {
        "filename": "faq_template.txt",
        "description": "文本模板：Q:/A: 必填，S: 相似问，C: 分类；条目间用 --- 分隔",
        "media_type": "text/plain; charset=utf-8",
    },
    "xls": {
        "filename": "faq_template.xls",
        "description": "Excel 97-2003 模板（下载后单元格内填 Q:/A: 内容）",
        "media_type": "application/vnd.ms-excel",
    },
    "xlsx": {
        "filename": "faq_template.xlsx",
        "description": "Excel 模板（单元格内填 Q:/A:，导入由 Tika 统一识别）",
        "media_type": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    },
}

_SUFFIX_MAP = {
    # 原有 5 种
    ".json": "json",
    ".csv": "csv",
    ".txt": "txt",
    ".text": "txt",
    ".xls": "xls",
    ".xlsx": "xlsx",
    # Tika 兜底新增：只要后缀在这里，就会用 Tika 抽文本然后走 Q/A 规则匹配
    ".pdf": "pdf",
    ".doc": "doc",
    ".docx": "docx",
    ".rtf": "rtf",
    ".odt": "odt",
    ".ppt": "ppt",
    ".pptx": "pptx",
    ".odp": "odp",
    ".html": "html",
    ".htm": "htm",
    ".md": "md",
    ".markdown": "markdown",
    ".eml": "eml",
}

_QA_PREFIX = re.compile(
    r"^(?:(Q|A|S|C|问|答|相似|分类)\s*[:：]\s*)(.*)$",
    re.IGNORECASE,
)


def list_template_meta() -> list[dict[str, str]]:
    return [
        {
            "format": fmt,
            "filename": _FORMAT_META[fmt]["filename"],
            "description": _FORMAT_META[fmt]["description"],
        }
        for fmt in TEMPLATE_FORMATS
    ]


def detect_format(name: str) -> str | None:
    suffix = Path(name or "").suffix.lower()
    return _SUFFIX_MAP.get(suffix)


def remap_entry_keys(item: dict[str, Any]) -> dict[str, Any]:
    """将中英文表头 / 字段名统一映射为内部英文键。"""
    return {_norm_header(str(k)): v for k, v in item.items() if k is not None}


def normalize_entry(item: dict[str, Any]) -> dict[str, Any] | None:
    item = remap_entry_keys(item)
    question = str(item.get("question") or "").strip()
    answer = str(item.get("answer") or "").strip()
    if not question or not answer:
        return None
    similar_raw = item.get("similar") or []
    if isinstance(similar_raw, str):
        similar = [
            p.strip()
            for p in re.split(r"[|；;]+", similar_raw)
            if p.strip()
        ]
    elif isinstance(similar_raw, list):
        similar = [str(s).strip() for s in similar_raw if str(s).strip()]
    else:
        similar = []
    out: dict[str, Any] = {
        "question": question,
        "answer": answer,
        "similar": similar,
        "category": str(item.get("category") or "").strip(),
    }
    faq_id = str(item.get("id") or "").strip()
    if faq_id:
        out["id"] = faq_id
    return out


def normalize_entries(raw_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        norm = normalize_entry(item)
        if norm:
            out.append(norm)
    return out


def _field_docs_json() -> list[dict[str, str]]:
    """JSON 模板内嵌的字段说明。"""
    return [
        {
            "字段名": d["field"],
            "中文表头": d["zh"],
            "是否必填": d["required"],
            "含义": d["meaning"],
        }
        for d in FIELD_DOCS
    ]


def parse_faq_bytes(
    data: bytes,
    fmt: str,
    *,
    filename_hint: str = "",
    settings: Settings | None = None,
) -> list[dict[str, Any]]:
    """**所有格式统一入口**：先用 Tika 抽纯文本，再按 Q/A/S/C 规则匹配 FAQ。

    Args:
      data: 上传文件字节
      fmt: 由 ``detect_format(文件名)`` 返回的格式标识（仅作校验；实际解析不依赖它）
      filename_hint: 原文件名，便于日志追踪
      settings: Settings 实例，用于取 TIKA_URL/超时/开关。未传则用 Settings.from_env()

    Returns:
      归一化后的 FAQ list[dict]，每一条至少含 question/answer；
      Tika 连不上 / 抽出来空 / 没有 Q/A 规则匹配 都会返回空列表，
      由上层接口返回 422 提示用户。
    """
    fmt = (fmt or "").strip().lower()
    if fmt not in SUPPORTED_FORMATS:
        raise ValueError(
            f"不支持的格式：{fmt}。可识别扩展名："
            + " / ".join(sorted({f".{f}" for f in _SUFFIX_MAP.keys()}))
        )
    if not data:
        return []
    settings = settings or Settings.from_env()
    text, _meta = extract_text_from_bytes(data, filename_hint, settings)
    if not text:
        # 降级兜底：UTF-8/GBK 强行 decode（防止用户传纯 txt 但 Tika 未启动）
        for enc in ("utf-8-sig", "utf-8", "gbk"):
            try:
                text = data.decode(enc).strip()
                if text:
                    break
            except UnicodeDecodeError:
                continue
    if not text:
        return []
    return _parse_text_content(text)


def parse_faq_path(
    path: Path,
    *,
    settings: Settings | None = None,
) -> list[dict[str, Any]]:
    path = path.resolve()
    if not path.exists():
        raise FileNotFoundError(f"FAQ 文件不存在: {path}")
    fmt = detect_format(path.name)
    if not fmt:
        raise ValueError(
            f"无法识别文件格式：{path.name}。支持扩展名："
            + " / ".join(sorted(_SUFFIX_MAP.keys()))
        )
    settings = settings or Settings.from_env()
    data = path.read_bytes()
    return parse_faq_bytes(data, fmt, filename_hint=path.name, settings=settings)


def parse_faq_url(
    url: str,
    *,
    settings: Settings | None = None,
) -> list[dict[str, Any]]:
    url = (url or "").strip()
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        raise ValueError("url 仅支持 http:// 或 https://")
    if not parsed.netloc:
        raise ValueError("url 无效")
    req = Request(url, headers={"User-Agent": "replykit/0.1", "Accept": "*/*"})
    try:
        with urlopen(req, timeout=30) as resp:
            body = resp.read()
    except HTTPError as exc:
        raise ValueError(f"拉取 FAQ URL 失败: HTTP {exc.code}") from exc
    except URLError as exc:
        raise ValueError(f"拉取 FAQ URL 失败: {exc.reason}") from exc
    fmt = detect_format(parsed.path) or "txt"
    settings = settings or Settings.from_env()
    return parse_faq_bytes(body, fmt, filename_hint=parsed.path, settings=settings)


def build_template_file(fmt: str) -> tuple[bytes, str, str]:
    """Return (content, filename, media_type). 仅限 TEMPLATE_FORMATS。"""
    fmt = (fmt or "").strip().lower()
    if fmt not in TEMPLATE_FORMATS:
        raise ValueError(
            f"未知模板格式：{fmt}，模板仅提供：{', '.join(TEMPLATE_FORMATS)}"
        )
    meta = _FORMAT_META[fmt]
    if fmt == "json":
        content = _render_json(SAMPLE_ENTRIES)
    elif fmt == "csv":
        content = _render_csv(SAMPLE_ENTRIES)
    elif fmt == "txt":
        content = _render_txt(SAMPLE_ENTRIES)
    elif fmt == "xls":
        content = _render_xls(SAMPLE_ENTRIES)
    else:
        content = _render_xlsx(SAMPLE_ENTRIES)
    return content, meta["filename"], meta["media_type"]


def _parse_text_content(text: str) -> list[dict[str, Any]]:
    """通用纯文本 FAQ 匹配：Q:/A:/S:/C: 前缀规则 + 分段分隔符。

    Tika 抽出来的 PDF / Excel / DOCX 纯文本都会进入这里。
    为兼容旧 Excel/JSON 中直接写作「标准问: xxx」「答案: yyy」的一行行内容，
    同样支持中文关键字 + 冒号 前缀（见 _QA_PREFIX 正则）。
    """
    entries: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    answer_lines: list[str] = []
    mode: str | None = None

    def flush() -> None:
        nonlocal current, answer_lines, mode
        if current is None:
            return
        if answer_lines:
            current["answer"] = "\n".join(answer_lines).strip()
        norm = normalize_entry(current)
        if norm:
            entries.append(norm)
        current = None
        answer_lines = []
        mode = None

    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        stripped = line.strip()
        if stripped in {"---", "***", "===", ""}:
            if stripped in {"---", "***", "==="}:
                flush()
            elif mode == "A" and answer_lines:
                answer_lines.append("")
            continue

        m = _QA_PREFIX.match(stripped)
        if m:
            tag, rest = m.group(1), m.group(2)
            tag_l = tag.lower()
            if tag_l in {"q", "问"}:
                flush()
                current = {
                    "question": rest.strip(),
                    "answer": "",
                    "similar": [],
                    "category": "",
                }
                answer_lines = []
                mode = "Q"
            elif tag_l in {"a", "答"}:
                if current is None:
                    current = {
                        "question": "",
                        "answer": "",
                        "similar": [],
                        "category": "",
                    }
                answer_lines = [rest] if rest.strip() else []
                mode = "A"
            elif tag_l in {"s", "相似"}:
                if current is None:
                    continue
                if rest.strip():
                    current.setdefault("similar", []).append(rest.strip())
                mode = "S"
            elif tag_l in {"c", "分类"}:
                if current is None:
                    continue
                current["category"] = rest.strip()
                mode = "C"
            continue

        if mode == "A" and current is not None:
            answer_lines.append(line)
        elif mode == "Q" and current is not None and stripped:
            current["question"] = f"{current.get('question', '')} {stripped}".strip()

    flush()
    return entries


def _render_json(entries: list[dict[str, Any]]) -> bytes:
    payload = {
        "字段说明": _field_docs_json(),
        "faqs": entries,
    }
    return (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8")


def _render_csv(entries: list[dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=TABLE_HEADERS_ZH,
        lineterminator="\n",
    )
    writer.writeheader()
    for item in entries:
        writer.writerow(_entry_to_zh_row(item))
    # BOM 方便 Excel 打开中文
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def _entry_to_zh_row(item: dict[str, Any]) -> dict[str, str]:
    return {
        "编号": str(item.get("id") or ""),
        "分类": str(item.get("category") or ""),
        "标准问": str(item.get("question") or ""),
        "答案": str(item.get("answer") or ""),
        "相似问": "|".join(item.get("similar") or []),
    }


def _append_field_docs_sheet_xlsx(wb: Any) -> None:
    ws = wb.create_sheet("字段说明", 0)
    ws.append(["表头（中文）", "内部字段名", "是否必填", "含义"])
    for d in FIELD_DOCS:
        ws.append([d["zh"], d["field"], d["required"], d["meaning"]])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 72


def _append_field_docs_sheet_xls(book: Any) -> None:
    sheet = book.add_sheet("字段说明")
    headers = ["表头（中文）", "内部字段名", "是否必填", "含义"]
    for c, h in enumerate(headers):
        sheet.write(0, c, h)
    for r, d in enumerate(FIELD_DOCS, start=1):
        sheet.write(r, 0, d["zh"])
        sheet.write(r, 1, d["field"])
        sheet.write(r, 2, d["required"])
        sheet.write(r, 3, d["meaning"])


def _render_txt(entries: list[dict[str, Any]]) -> bytes:
    parts: list[str] = [
        "# FAQ 文本模板：每条以 Q:/A: 开头；S: 相似问（可多行）；C: 分类；条目之间用 --- 分隔",
        "# 字段说明：Q=标准问(必填) A=答案(必填) S=相似问(可选,可多行) C=分类(可选)",
        "",
    ]
    for i, item in enumerate(entries):
        if i:
            parts.append("---")
            parts.append("")
        parts.append(f"Q: {item.get('question', '')}")
        parts.append(f"A: {item.get('answer', '')}")
        for s in item.get("similar") or []:
            parts.append(f"S: {s}")
        if item.get("category"):
            parts.append(f"C: {item['category']}")
        parts.append("")
    return ("\n".join(parts)).encode("utf-8")


def _render_xlsx(entries: list[dict[str, Any]]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    # 先写数据表，再把「字段说明」插到最前，打开文件先看到说明
    ws = wb.active
    ws.title = "FAQ数据"
    ws.append(list(TABLE_HEADERS_ZH))
    for item in entries:
        row = _entry_to_zh_row(item)
        ws.append([row[h] for h in TABLE_HEADERS_ZH])
    for col, width in enumerate([12, 12, 28, 48, 36], start=1):
        ws.column_dimensions[chr(64 + col)].width = width
    _append_field_docs_sheet_xlsx(wb)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _render_xls(entries: list[dict[str, Any]]) -> bytes:
    import xlwt

    book = xlwt.Workbook(encoding="utf-8")
    _append_field_docs_sheet_xls(book)
    sheet = book.add_sheet("FAQ数据")
    for c, h in enumerate(TABLE_HEADERS_ZH):
        sheet.write(0, c, h)
    for r, item in enumerate(entries, start=1):
        row = _entry_to_zh_row(item)
        for c, h in enumerate(TABLE_HEADERS_ZH):
            sheet.write(r, c, row[h])
    bio = io.BytesIO()
    book.save(bio)
    return bio.getvalue()
